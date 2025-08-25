import pandas as pd
import openpyxl

def csv_to_excel_with_formulas(csv_file, excel_file):
    try:
        # Load existing workbook, keeping formulas
        wb = openpyxl.load_workbook(excel_file, keep_links=True)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Read CSV and skip the first column
    df = pd.read_csv(csv_file, dtype=str).iloc[:, 1:]

    # Ensure "Indexdata" sheet exists
    if "Indexdata" in wb.sheetnames:
        ws = wb["Indexdata"]
        ws.delete_rows(2, ws.max_row)  # Clear data but keep formulas
    else:
        ws = wb.create_sheet(title="Indexdata")

    # Write headers
    for col_idx, header in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data, preserving formulas and handling numbers
    for r_idx, row in enumerate(df.values, start=2):  
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, str) and value.startswith("="):
                ws.cell(row=r_idx, column=c_idx, value=value)  # Keep formulas
            else:
                try:
                    ws.cell(row=r_idx, column=c_idx, value=float(value))  # Convert numbers
                except ValueError:
                    ws.cell(row=r_idx, column=c_idx, value=value)  # Keep text

    # Save workbook
    wb.save(excel_file)
    print(f"Updated {excel_file}, preserving all sheets and formulas.")

# Example usage
csv_to_excel_with_formulas("NIFTY  indices value.csv", "Sheet\\DAILYREPORT.xlsm")
