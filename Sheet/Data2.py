import pandas as pd
from openpyxl import load_workbook

# File paths
excel_path = "Sheet\DAILY REPORT.xlsm"
csv_paths = [
    "900 and Above.csv",
    "200-899.csv",
    "0-199.csv"
]

# Load and combine CSV data, excluding the first column
combined_data = pd.concat([pd.read_csv(csv).iloc[:, 1:] for csv in csv_paths], ignore_index=True)

# Load the existing Excel file
wb = load_workbook(excel_path, data_only=False)  # Load formulas

# Write merged data to "Data2" sheet
if "Data 2" in wb.sheetnames:
    ws = wb["Data 2"]
    
    # Store formulas in a dictionary { (row, col): formula }
    formulas = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):  # Check if it's a formula
                formulas[(cell.row, cell.column)] = cell.value
    
    # Clear existing data except header
    ws.delete_rows(2, ws.max_row)
    
    # Insert new data
    for r_idx, row in enumerate(combined_data.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Restore formulas
    for (row, col), formula in formulas.items():
        ws.cell(row=row, column=col, value=formula)
else:
    print("Sheet 'Data2' not found!")

# Save the updated file
wb.save(excel_path)
print("Excel file updated successfully, preserving formulas!")
