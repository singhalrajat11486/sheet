from openpyxl import load_workbook

def copy_paste_values(file_path, sheet_name):
    # Load workbook in data-only mode to get computed values
    wb_read = load_workbook(filename=file_path, data_only=True, keep_vba=True)  
    ws_read = wb_read[sheet_name]

    # Load workbook normally to allow editing
    wb_write = load_workbook(filename=file_path, keep_vba=True)
    ws_write = wb_write[sheet_name]

    # Define source and destination ranges
    copy_ranges = [("D3:D10", "E3"), ("D13:D17", "E13"), ("D20:D43", "E20"),
                   ("I3:I10", "J3"), ("I13:I17", "J13"), ("I20:I43", "J20")]

    for copy_range, paste_start in copy_ranges:
        # Extract source and destination ranges
        start_cell, end_cell = copy_range.split(":")
        start_col, start_row = start_cell[0], int(start_cell[1:])
        end_row = int(end_cell[1:])

        paste_col, paste_row = paste_start[0], int(paste_start[1:])

        # Copy values only (ignoring formulas)
        for i, row in enumerate(range(start_row, end_row + 1)):
            src_cell = ws_read[f"{start_col}{row}"]  # Read only computed value
            dest_cell = ws_write[f"{paste_col}{paste_row + i}"]

            dest_cell.value = src_cell.value  # Copy only the value, ignoring formulas

    # Save workbook with values pasted
    wb_write.save(filename=file_path)
    print("✅ Values copied successfully! No formulas were copied.")

# Example usage
copy_paste_values("Sheet/DAILY REPORT.xlsm", "Breadth")
